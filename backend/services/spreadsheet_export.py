"""Formatted .xlsx exports of the read-only stats (D.3).

Parent / investigation-facing playing-time summaries. DELIBERATELY excludes skill
ratings and any internal fairness/rotation settings — the sheet shows only who
played, how much (matches, slots, minutes), goals, and a position spread.
"""
from __future__ import annotations

import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]  # ships no type stubs
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from sqlmodel import Session

from backend.db.models import SquadDB, TournamentDB
from backend.services import analytics

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADERS = ["Player", "Matches", "Slots", "Minutes", "Goals", "GK", "DEF", "MID", "FWD"]


def _sanitize(name: str) -> str:
    """Filesystem-safe filename fragment (letters/digits/space/-/_)."""
    keep = "".join(c if (c.isalnum() or c in " -_") else " " for c in name)
    return " ".join(keep.split()) or "Gaffer"


def _row_from_stat(s: dict[str, Any]) -> list[Any]:
    pos = s.get("positions") or {}
    return [
        s["name"], s.get("matches_available", 0), s["slots_played"],
        s.get("minutes", 0), s.get("goals", 0),
        pos.get("GK", 0), pos.get("DEF", 0), pos.get("MID", 0), pos.get("FWD", 0),
    ]


def _build_workbook(title: str, subtitle: str, rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # a fresh Workbook always has an active sheet
    ws.title = "Stats"

    ncols = len(_HEADERS)
    center = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor="1A5276")
    header_font = Font(bold=True, color="FFFFFF")

    # Title + subtitle (merged across the table width).
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.append([subtitle])
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(_HEADERS)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        if col > 1:
            cell.alignment = center

    total_slots = total_minutes = total_goals = 0
    for r in sorted(rows, key=lambda x: str(x[0]).lower()):
        ws.append(r)
        total_slots += r[2]
        total_minutes += r[3]
        total_goals += r[4]
        for col in range(2, ncols + 1):
            ws.cell(row=ws.max_row, column=col).alignment = center

    ws.append(["TOTAL", "—", total_slots, total_minutes, total_goals, "", "", "", ""])
    for col in range(1, ncols + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
        if col > 1:
            cell.alignment = center

    ws.append([])
    ws.append(["Minutes are based on the recorded rotation plan (planned playing time)."])

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.column_dimensions["A"].width = 22
    for col in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _team_name(session: Session, squad_id: int) -> str:
    squad = session.get(SquadDB, squad_id)
    return (squad.team_name or squad.name if squad else None) or "Level"


def _today() -> str:
    return datetime.date.today().isoformat()


def season_workbook(session: Session, squad_id: int) -> tuple[bytes, str]:
    team = _team_name(session, squad_id)
    rows = [_row_from_stat(s) for s in analytics.season_stats(session, squad_id)]
    data = _build_workbook(team, f"Season summary · generated {_today()}", rows)
    return data, f"{_sanitize(team)} season stats {_today()}.xlsx"


def tournament_workbook(session: Session, squad_id: int, tournament_id: int) -> tuple[bytes, str]:
    t = session.get(TournamentDB, tournament_id)
    name = (t.name if t else None) or "Tournament"
    stats = analytics.tournament_stats(session, squad_id, tournament_id)["players"]
    rows = [_row_from_stat(s) for s in stats]
    data = _build_workbook(name, f"Tournament summary · generated {_today()}", rows)
    return data, f"{_sanitize(name)} stats {_today()}.xlsx"


def all_tournaments_workbook(session: Session, squad_id: int) -> tuple[bytes, str]:
    team = _team_name(session, squad_id)
    rows = [_row_from_stat(s) for s in analytics.all_tournament_stats(session, squad_id)["players"]]
    data = _build_workbook(
        f"{team} — All tournaments",
        f"All tournament matches · generated {_today()}",
        rows,
    )
    return data, f"{_sanitize(team)} all tournament stats {_today()}.xlsx"
